import time
import os
import requests
from collections import Counter
import spacy
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from bs4 import BeautifulSoup
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openai import OpenAI
import pandas as pd
from datetime import datetime
from selenium.webdriver.chrome.options import Options
import traceback

def obtenerCodigoWeb(url):
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--disable-cookies") 
    #Se evitan las coockies al acceder a la url pasada por parámetro
    driver = webdriver.Chrome(options=chrome_options)
    driver.get(url)
    #se crea la variable "wait" en la que indicamos que el buscador debe esperar 10 segundos
    wait = WebDriverWait(driver, 10)
    #Se indica al buscador que espere los 10 segundos a no ser que se haya cargado todo el body
    wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
    #Se obtiene todo el código html de la url
    html_content = driver.page_source
    #Cerramos la sesión del navegador
    driver.quit()
    return html_content

def procesarRespuesta(texto,url,eissn,journal,web):
    datosExcel=[]
    #Se splittea para separar la información de cada miembro editorial
    lineas = texto.split("#")
    if lineas[-1] == '':     # Elimina el último elemento vacío si existe
        lineas.pop() 
    for linea in lineas:
        #Se splittea la información entre la línea para separar los datos obtenidos por chatGPT
        datos=linea.split("/")
        #Datos[0] = nombre       #Datos[1] = País del usuario del que se almacenan los datos        #Datos[2] = Lugar donde trabaja
        if datos[0]!="NA" and datos[0]!=" NA" and datos[0] != "n" and "name" not in datos[0] and datos[0]!= "Name: NA" and len(datos)>=3 and datos[0]!= " " and datos[0]!= "ACM" and "editor" not in datos[0]:             #En ocasiones el sistema devuelve datos extraños con todos los valores "NA", con este if se eliminan estos casos
            #Se evitan las "," que se añaden en algunas revistas entre país y lugar de trabajo, se reemplazan por un vacío
            #se evitan algunas posibilidades que devuelve ChatGPT para asegurar que los datos se devuelven como deben
            datos[0]=datos[0].replace(",","")
            datos[1]=datos[1].replace(",","")
            datos[2]=datos[2].replace(",","")
            orcid = "NA"
            #En caso de que ambos datos que forman la afiliación estén vacíos
            if datos[1] =="NA" and datos[2]=="NA":
                if web:
                    urlOrcid = obtenerVariable("BASIC_URL") + datos[0].replace(" ","%20")
                    informacion = accederORCID(urlOrcid)
                    afiliacion = informacion.split("!@!")[0]
                    orcid = informacion.split("!@!")[1]
                else:
                    afiliacion = "NA"
            if datos[1] != "NA" and datos[2]=="NA": #En caso de que se sepa el país pero no donde trabaja
                afiliacion = datos[1]
            if datos[1] == "NA" and datos[2]!="NA": #En caso de que no se sepa el país pero si donde trabaja
                afiliacion = datos[2]
            if datos[1] != "NA" and datos[2]!="NA": #En caso de que se sepan ambos datos
                afiliacion = datos[1] + " " + datos[2]
            #Se crea la línea del excel, se añade al array y se vuelve a la siguiente lista
            lineaExcel = datos[0] + "," + orcid + "," + afiliacion  +",NA," + journal + ",NA," + eissn + ","
            lineaExcel = lineaExcel + datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "," +url
            datosExcel.append(lineaExcel)
    return datosExcel

def obtenerVariable(variable):
    #Se carga el archivo .env
    load_dotenv("variables.env")
    #Se accede al archivo y se retorna el valor de la variable 
    var = os.getenv(variable)
    return var


def completarConversacion(client, bodyContent, prompt):
    return client.chat.completions.create(
        model="gpt-3.5-turbo",
        messages=[
            {
                "role": "system",
                "content": prompt
            },
            {
                "role": "user",
                "content": bodyContent
            }
        ],
        temperature=0.7,
        max_tokens=None,
        top_p=0.7
    )

def accederORCID(url):
    driver = webdriver.Chrome() #Se accede a Google chrome
    driver.get(url) #Se añade la url en el buscador de Google chrome
    wait = WebDriverWait(driver, 10) #Se espera 10 segundos a conectar a lka página web
    boton = wait.until(EC.presence_of_element_located((By.ID, "onetrust-reject-all-handler")))
    #Se pulsa el boton "Boton" con el que, posteriormente, podremos acceder a la información
    boton.click()
    time.sleep(2) #pulsamos el boton y realizamos una espera de 2 segundos para evitar que se nos identifique como bots
    elemento = driver.find_elements(By.CSS_SELECTOR, "tr.ng-star-inserted") 
    #comprobamos si existe la tabla en la que se almacena la informacion
    if not elemento: #Si no hay tabla, ambos datos a "NA"
        afiliacion = "NA"
        orcid = "NA"
    else: #En caso de que haya información esperamos a que cargue y se extrae la información
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "tr.ng-star-inserted")))
        html_content = driver.page_source
        soup = BeautifulSoup(html_content, 'html.parser')
        #Se extrae la primera fila de información
        trList = soup.find_all('tr', class_='ng-star-inserted')[:1] 
        for tr in trList: #Accedemos a cada columna d ela primera fila 
            td_list = tr.find_all('td')  
            ultimo_td = td_list[-1]  #Obtenemos la afiliación que está ubicada en la última columna
            primer_td = td_list[0] #Obtenemos el ORCID que está en la primera columna 
            #Extraemos el texto de ambas informaciones y comprobamos que haya datos
            datos = ultimo_td.get_text().replace('\n', '').strip()
            if(datos == ""):
                afiliacion = "NA"
            else:
                afiliacion = datos
            datosORCID = primer_td.get_text().replace('\n', '').strip()
            if (datosORCID == ""):
                orcid = "NA"
            else:
                orcid = datosORCID
    #Salimos de Google Chrome y devolvemos la información
    driver.quit()
    informacion = afiliacion + "!@!" + orcid
    return informacion

def comenzarEstudio(web):
    if  web:  #Se obtiene el nombre del excel de destino
        excel = obtenerVariable("ExcelMiembrosAfiliacion")
    else:
        excel = obtenerVariable("EXCEL_MIEMBROS")
    excel = "muestra_openeditors_copia.xlsx"
    wb = load_workbook(excel)     #se carga el excel y se abre
    ws = wb.active
    #Accedemos al excel con los datos de destino y extraemos los campos necesarios
    dataframe = pd.read_excel("jcr_computer_science_journals_pfg_resultados.xlsx") 
    urls = dataframe[["URL", "ISSN","Journal name"]][0:1] 
    try:     
        for  index,fila in urls.iterrows(): #Para cada una de las filas
            #url = fila["URL"]
            url = "https://journals.sagepub.com/editorial-board/adba"
            #issn =  fila["ISSN"]
            issn= "1059-7123"
            if issn == "NO":
                issn = "NA"
            #nombre = fila["Journal name"]
            nombre = "Adaptive Behavior"
            #Se extrae el codigo html de la página web
            html = obtenerCodigoWeb(url) 
            #Se extrae el contenido del body del html
            body = obtenerTexto(html)
            #En caso de que el body sea demasiado grande, se divide en partes con el tamaño de tokens máximos de chatGPT 
            partes_texto = dividirTexto(body,16000)
            #Instanciamos de la clase OpenAI
            client = OpenAI(api_key=obtenerVariable("OPENAI_API_KEY")) 
            #Usamos el objeto instanciado y hacemos la consulta a ChatGPT en partes de tamaño óptimo para ChatGPT
            respuesta= ""
            for texto in (partes_texto):
                respuestaGPT = completarConversacion(client, texto, obtenerVariable("PROMPTURLTFG"))
                mensaje = respuestaGPT.choices[0].message.content
                respuesta = respuesta + mensaje
            #Procesamos la respuesta de todos los textos de ChatGPT
            textoProcesado = procesarRespuesta(respuesta,url,issn,nombre,web)
            #Almacenamos en el excel los datos procesados de ChatGPT
            añadirEnExcel(ws,wb,textoProcesado,excel)
    except Exception as e:
        print(f"Error al procesar la url: " + url)
        print(traceback.format_exc())
    finally:
        wb.save(excel) 
        
def obtenerTexto(html):
    soup = BeautifulSoup(html, 'html.parser')
    body = soup.find('body').get_text()
    return body
        
def añadirEnExcel(ws,wb,lineas,excel):
    for linea in lineas:
        filaVacia = ws.max_row + 1
        ws.cell(row=filaVacia, column=1, value=linea.replace("\n", ""))
    filaVacia = ws.max_row + 1
    ws.cell(row=filaVacia, column=1, value="--------------------------------------------")
    wb.save(excel) 


def dividirTexto(texto, max_token):
    #Se crea el array donde se almacenarán los textos enviados a ChatGPT
    respuesta = []
    #se revisa la cantidad de carácteres del texto
    if len(texto) <= max_token:
    #Si los caracteres son menores a 16000 se envía todo el texto a chatGPT
        respuesta.append(texto)
    else:
    #Si hay más de 16000 caracteres se divide en varios
        token_actual = 0
        texto_nuevo = ""
        palabras = texto.split()  # Dividimos el texto en palabras
        for palabra in palabras:
    #Si todas las palabras añadidas + la nueva palabra que se va a añadir al texto es < a 16000
            if  len(palabra) + len(texto_nuevo) <= max_token:
                if texto_nuevo: #separamos las palabras y añadimos la nueva
                    texto_nuevo += " "
                texto_nuevo += palabra
                token_actual += len(palabra)
            else: # Si la nueva palabra añadida sumase más de 16000 caracteres, se crea un nuevo texto
                respuesta.append(texto_nuevo)
                texto_nuevo = palabra
                token_actual = len(palabra)
        if texto_nuevo:
            respuesta.append(texto_nuevo)
    return respuesta #Se devuelve el array con todos los textos
            